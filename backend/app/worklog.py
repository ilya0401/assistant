from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .config import settings

HEADERS = ["#", "Номер задачи", "Дата выполнения задачи", "Время на задачу", "Действия по задаче", "Добавлено"]
HEADER_COLOR = "2D3748"


def _workbook_path() -> Path:
    return Path(settings.data_dir) / "worklog.xlsx"


def _init_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Worklog"
    ws.append(HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 22

    wb.save(_workbook_path())
    return wb


def save_entry(task: str, date: str, time_spent: str, description: str) -> int:
    path = _workbook_path()
    wb = openpyxl.load_workbook(path) if path.exists() else _init_workbook()
    ws = wb.active

    entry_id = ws.max_row  # row 1 = header, so first entry gets id=1
    ws.append([entry_id, task, date, time_spent, description, datetime.now().strftime("%Y-%m-%d %H:%M")])
    wb.save(path)
    return entry_id


def get_entries(limit: int = 20) -> list[dict]:
    path = _workbook_path()
    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    result = []
    for row in rows[-limit:]:
        if row[0] is not None:
            result.append({
                "id": row[0],
                "task": row[1],
                "date": str(row[2]) if row[2] else "",
                "time_spent": row[3],
                "description": row[4],
                "added": str(row[5]) if row[5] else "",
            })
    return list(reversed(result))
